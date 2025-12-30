import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# 输出文件名定义
OUTPUT_FILE_485 = 'DR485-Kinetics.xlsx'
OUTPUT_FILE_420 = 'DR420-Kinetics.xlsx'
OUTPUT_FILE_DF485 = 'DF485-Kinetics.xlsx'
OUTPUT_FILE_DF420 = 'DF420-Kinetics.xlsx'
OUTPUT_FILE_DR = 'DR.xlsx'

MODULES = {
    'DR485': (41, 46, 48, OUTPUT_FILE_485),
    'DR420': (61, 66, 68, OUTPUT_FILE_420),
    'DF485': (126, 131, 132, OUTPUT_FILE_DF485),
    'DF420': (145, 150, 151, OUTPUT_FILE_DF420),
}

def generate_wells():
    return [f'{r}{c:02d}' for r in 'ABCDEFGH' for c in range(1, 13)]

def extract_module(df, time_start, time_end, dr_row, label):
    block = df.iloc[time_start:time_end + 1, 2:98].T.reset_index(drop=True)
    block.columns = [f'T{i+1}' for i in range(6)]
    dr = df.iloc[dr_row, 2:98].T.reset_index(drop=True).rename(label)
    au = df.iloc[50, 2:98].T.reset_index(drop=True).rename('AU')
    merged = pd.concat([block, dr, au], axis=1)
    return merged

def enrich_dataframe(df):
    df.insert(0, 'Plate', df['Source'].str.split('-').str[0])
    df['Treatment'] = df['Source'].str.split('-').str[1]

    wells_96 = generate_wells()
    df['Sample'] = None

    for source, idxs in df.groupby('Source').groups.items():
        plate = source.split('-')[0]
        wells = wells_96[:len(idxs)]
        df.loc[idxs, 'Sample'] = [f"{plate}-{w}" for w in wells]

    cols = df.columns.tolist()
    cols.insert(1, cols.pop(cols.index('Sample')))
    cols.append(cols.pop(cols.index('Treatment')))
    return df[cols]

def process_and_save(data, label, output_file, on_step=None):
    final_df = pd.concat(data, ignore_index=True)
    numeric_cols = final_df.select_dtypes(include='number').columns.tolist()
    final_df[numeric_cols] = final_df[numeric_cols].astype(float).round(2)

    final_df = enrich_dataframe(final_df)

    # ✅ 添加 ID 列
    final_df.insert(0, 'ID', range(1, len(final_df) + 1))

    final_df.to_excel(output_file, index=False)
    if on_step: on_step()

    wb = load_workbook(output_file)
    ws = wb.active
    col_index_map = {col: idx + 1 for idx, col in enumerate(final_df.columns) if col in numeric_cols}

    for row in ws.iter_rows(min_row=2):
        for col_idx in col_index_map.values():
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
    wb.save(output_file)
    if on_step: on_step()

    apply_column_gradient(output_file, target_columns=['AU'], mode='two-color',
                          start_color='FFFFFF', end_color='4169E1')
    if on_step: on_step()

    apply_column_gradient(output_file, target_columns=[f'T{i+1}' for i in range(6)],
                          mode='three-color')
    if on_step: on_step()

    round_numeric_cells_in_excel(output_file)
    if on_step: on_step()

def apply_column_gradient(file_path, target_columns,
                          mode='three-color', start_color='00FF00',
                          mid_color='FFFFFF', end_color='FF69B4'):
    wb = load_workbook(file_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    for col_name in target_columns:
        if col_name not in header:
            continue
        col_index = header.index(col_name) + 1
        col_letter = ws.cell(row=1, column=col_index).column_letter
        col_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        if mode == 'two-color':
            rule = ColorScaleRule(start_type='min', start_color=start_color,
                                  end_type='max', end_color=end_color)
        else:
            rule = ColorScaleRule(start_type='min', start_color=start_color,
                                  mid_type='num', mid_value=0, mid_color=mid_color,
                                  end_type='max', end_color=end_color)

        ws.conditional_formatting.add(col_range, rule)

    wb.save(file_path)

def round_numeric_cells_in_excel(file_path, decimal_places=2):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.value = round(cell.value, decimal_places)
                cell.number_format = f'0.{"0" * decimal_places}'
    wb.save(file_path)

def handle_processing(source_file, on_step=None):
    excel = pd.ExcelFile(source_file)
    data_dr = []
    data_modules = {label: [] for label in MODULES}
    output_folder = os.path.dirname(source_file)

    for sheet in excel.sheet_names:
        df = excel.parse(sheet_name=sheet, header=None)

        for label, (start, end, dr_row, _) in MODULES.items():
            module_data = extract_module(df, start, end, dr_row, label)
            module_data['Source'] = sheet
            data_modules[label].append(module_data)
            if on_step:
                for _ in range(8):  # T1–T6 + DR + AU
                    on_step()

        dr_vals = {
            'DR485': df.iloc[48, 2:98].T.reset_index(drop=True).rename('DR485'),
            'DR420': df.iloc[68, 2:98].T.reset_index(drop=True).rename('DR420'),
            'DF485': df.iloc[132, 2:98].T.reset_index(drop=True).rename('DF485'),
            'DF420': df.iloc[151, 2:98].T.reset_index(drop=True).rename('DF420'),
            'AU': df.iloc[50, 2:98].T.reset_index(drop=True).rename('AU')
        }
        merged = pd.concat(list(dr_vals.values()), axis=1)
        merged['Source'] = sheet
        data_dr.append(merged)
        if on_step:
            for _ in range(6):  # 5 columns + Source
                on_step()

    output_files = []
    for label, (_, _, _, filename) in MODULES.items():
        output_path = os.path.join(output_folder, filename)
        process_and_save(data_modules[label], label, output_path, on_step)
        output_files.append(output_path)

    dr_output_path = os.path.join(output_folder, OUTPUT_FILE_DR)
    final_df = pd.concat(data_dr, ignore_index=True)
    numeric_cols = final_df.select_dtypes(include='number').columns.tolist()
    final_df[numeric_cols] = final_df[numeric_cols].astype(float).round(2)

    final_df = enrich_dataframe(final_df)

    # ✅ 添加 ID 列
    final_df.insert(0, 'ID', range(1, len(final_df) + 1))

    final_df.to_excel(dr_output_path, index=False)
    if on_step: on_step()

    wb = load_workbook(dr_output_path)
    ws = wb.active
    col_index_map = {col: idx + 1 for idx, col in enumerate(final_df.columns) if col in numeric_cols}
    for row in ws.iter_rows(min_row=2):
        for col_idx in col_index_map.values():
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
    wb.save(dr_output_path)
    if on_step: on_step()

    apply_column_gradient(dr_output_path, target_columns=['DR485', 'DR420', 'DF485', 'DF420'],
                          mode='three-color')
    if on_step: on_step()

    apply_column_gradient(dr_output_path, target_columns=['AU'], mode='two-color',
                          start_color='FFFFFF', end_color='4169E1')
    if on_step: on_step()

    round_numeric_cells_in_excel(dr_output_path)
    if on_step: on_step()

    output_files.append(dr_output_path)
    return output_files

def run_main(file_path, on_step=None):
    return handle_processing(file_path, on_step)
