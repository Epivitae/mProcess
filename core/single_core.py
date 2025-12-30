import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def run_main(filepath, on_step=None):
    kinetics_file = 'Kinetics.xlsx'
    dr_file = 'DR.xlsx'

    xl = pd.ExcelFile(filepath)
    wells = [r + f"{c:02d}" for r in 'ABCDEFGH' for c in range(1, 13)]
    data_frames = []

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name, header=None)
        block = df.iloc[13:23, 2:98].T.reset_index(drop=True)
        block.columns = [f'T{i+1}' for i in range(10)]

        dr = df.iloc[22, 2:98].T.reset_index(drop=True).rename('DR')
        au = df.iloc[13, 2:98].T.reset_index(drop=True).rename('AU')

        plate = sheet_name.split('-')[0]
        treatment = sheet_name.split('-')[1] if '-' in sheet_name else ''
        sample_ids = [f"{plate}-{w}" for w in wells]

        meta = pd.DataFrame({
            'Plate': plate,
            'Sample': sample_ids,
            'Source': sheet_name,
            'Treatment': treatment
        })

        merged = pd.concat([meta, block, dr, au], axis=1)
        data_frames.append(merged)

        if on_step:
            for _ in range(12):  # Plate, Sample, Source, Treatment + T1–T10 + DR + AU
                on_step()

    full_df = pd.concat(data_frames, ignore_index=True)
    full_df.to_excel(kinetics_file, index=False)
    if on_step: on_step()

    apply_shared_gradient(kinetics_file, ['T1','T2','T3','T4','T5','T6','DR'], on_step)
    apply_individual_gradient(kinetics_file, ['AU'], on_step)

    df_kinetics = pd.read_excel(kinetics_file)
    df_dr = df_kinetics[['Sample', 'Plate', 'Source', 'Treatment', 'DR', 'AU']].copy()
    df_dr.insert(0, 'ID', range(1, len(df_dr) + 1))
    df_dr.to_excel(dr_file, index=False)
    if on_step: on_step()

    apply_individual_gradient(dr_file, ['DR', 'AU'], on_step)

    return move_and_cleanup([kinetics_file, dr_file], filepath)

def apply_shared_gradient(file_path, columns, on_step=None):
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [str(cell.value).strip() for cell in ws[1]]
    df = pd.DataFrame(ws.iter_rows(min_row=2, values_only=True), columns=headers)

    cols = [col for col in columns if col in df.columns]
    if not cols:
        wb.save(file_path)
        return

    values = pd.concat([pd.to_numeric(df[col], errors='coerce') for col in cols], axis=1).values.flatten()
    valid = pd.Series(values).dropna()
    if valid.empty:
        wb.save(file_path)
        return

    min_val, max_val = valid.min(), valid.max()
    col_letters = [get_column_letter(headers.index(col) + 1) for col in cols]
    ref = f"{col_letters[0]}2:{col_letters[-1]}{ws.max_row}"
    rule = ColorScaleRule(
        start_type='num', start_value=min_val, start_color='00FF00',
        mid_type='num', mid_value=0, mid_color='FFFFFF',
        end_type='num', end_value=max_val, end_color='FF00FF'
    )
    ws.conditional_formatting.add(ref, rule)
    wb.save(file_path)
    if on_step: on_step()

def apply_individual_gradient(file_path, columns, on_step=None):
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [str(cell.value).strip() for cell in ws[1]]
    df = pd.DataFrame(ws.iter_rows(min_row=2, values_only=True), columns=headers)

    for col in columns:
        if col not in df.columns:
            continue
        values = pd.to_numeric(df[col], errors='coerce').dropna()
        if values.empty:
            continue
        min_val, max_val = values.min(), values.max()
        col_letter = get_column_letter(headers.index(col) + 1)
        ref = f"{col_letter}2:{col_letter}{ws.max_row}"
        rule = ColorScaleRule(
            start_type='num', start_value=min_val, start_color='00FF00',
            mid_type='num', mid_value=0, mid_color='FFFFFF',
            end_type='num', end_value=max_val, end_color='FF00FF'
        )
        ws.conditional_formatting.add(ref, rule)
        if on_step: on_step()
    wb.save(file_path)

def move_and_cleanup(files, source_path):
    target_dir = os.path.dirname(os.path.abspath(source_path))
    moved_files = []
    for file in files:
        target_path = os.path.join(target_dir, file)
        shutil.copy(file, target_path)
        os.remove(file)
        moved_files.append(target_path)
    print(f"✅ 文件已复制并清理：{target_dir}")
    return moved_files
